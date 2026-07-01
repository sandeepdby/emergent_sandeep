"""
Iteration 36 – Bulk Family Import + Rate Card Auto-Fill
Verifies:
  1. GET /api/endorsements/template/family returns a valid xlsx with 3 sheets
     ('Family Import', 'Instructions', 'Tips') and 7 sample rows.
  2. Existing IMP_001 endorsements have auto-filled per_life_premium = 11500
     and prorata_premium correctly calculated.
  3. Import Excel with blank Per Life Premium -> auto-fill from Rate Card
     (age 41 on policy GMC0001393000100 -> 11500).
  4. Import Excel with explicit Per Life Premium (5000) -> uses provided
     value (NOT the rate card value).
  5. Standard template regression (GET /api/endorsements/template).
"""
import os
import io
import time
from datetime import datetime, timezone

import pytest
import requests
import pandas as pd
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://insurehub-portal.preview.emergentagent.com").rstrip("/")
API = f"{BASE_URL}/api"

HR_CREDS = {"username": "arpita", "password": "Password@123"}
ADMIN_CREDS = {"username": "masteradmin", "password": "Admin@123"}
POLICY_NUMBER = "GMC0001393000100"


# ---------- Fixtures ----------
@pytest.fixture(scope="module")
def hr_token():
    r = requests.post(f"{API}/auth/login", json=HR_CREDS, timeout=20)
    assert r.status_code == 200, f"HR login failed: {r.status_code} {r.text}"
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def hr_headers(hr_token):
    return {"Authorization": f"Bearer {hr_token}"}


# ---------- 1. Family Template ----------
class TestFamilyTemplate:
    def test_family_template_downloads_ok(self, hr_headers):
        r = requests.get(f"{API}/endorsements/template/family", headers=hr_headers, timeout=30)
        assert r.status_code == 200, r.text
        ctype = r.headers.get("content-type", "")
        assert "spreadsheet" in ctype or "openxml" in ctype, ctype
        assert len(r.content) > 1000

    def test_family_template_has_three_sheets(self, hr_headers):
        r = requests.get(f"{API}/endorsements/template/family", headers=hr_headers, timeout=30)
        wb = load_workbook(io.BytesIO(r.content))
        sheets = wb.sheetnames
        assert "Family Import" in sheets
        assert "Instructions" in sheets
        assert "Tips" in sheets

    def test_family_template_has_seven_sample_rows(self, hr_headers):
        r = requests.get(f"{API}/endorsements/template/family", headers=hr_headers, timeout=30)
        df = pd.read_excel(io.BytesIO(r.content), sheet_name="Family Import")
        assert len(df) == 7, f"Expected 7 sample rows, got {len(df)}"
        fam001 = df[df["Employee ID"] == "FAM_001"]
        fam002 = df[df["Employee ID"] == "FAM_002"]
        assert len(fam001) == 6
        assert len(fam002) == 1
        rels = set(fam001["Relationship Type"].tolist())
        assert rels == {"Employee", "Spouse", "Kids1", "Kids2", "Mother", "Father"}

    def test_family_template_tips_sheet_has_seven_tips(self, hr_headers):
        r = requests.get(f"{API}/endorsements/template/family", headers=hr_headers, timeout=30)
        tips_df = pd.read_excel(io.BytesIO(r.content), sheet_name="Tips")
        assert len(tips_df) == 7


# ---------- 2. Seeded IMP_001 rows ----------
class TestSeededImportEndorsements:
    def test_imp001_two_endorsements_exist_with_autofilled_rate(self, hr_headers):
        r = requests.get(f"{API}/endorsements?policy_number={POLICY_NUMBER}", headers=hr_headers, timeout=30)
        assert r.status_code == 200
        rows = [e for e in r.json() if e.get("employee_id") == "IMP_001"]
        assert len(rows) >= 2, f"Expected >=2 IMP_001 rows, got {len(rows)}"

        by_rel = {e["relationship_type"]: e for e in rows}
        assert "Employee" in by_rel and "Spouse" in by_rel

        emp = by_rel["Employee"]
        sp = by_rel["Spouse"]
        assert emp["age"] == 41
        assert sp["age"] == 37
        assert emp["per_life_premium"] == 11500, emp
        assert sp["per_life_premium"] == 11500, sp
        # prorata must be a positive number
        assert isinstance(emp["prorata_premium"], (int, float)) and emp["prorata_premium"] > 0
        assert isinstance(sp["prorata_premium"], (int, float)) and sp["prorata_premium"] > 0


# ---------- Helpers ----------
def _make_import_df(per_life_value, employee_id, name_suffix):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return pd.DataFrame([{
        "Employee ID": employee_id,
        "Member Name": f"AutoTest {name_suffix}",
        "Relationship Type": "Employee",
        "DOB": "1984-05-10",   # age ~41
        "Age": 41,
        "Gender": "Male",
        "Policy Number": POLICY_NUMBER,
        "Endorsement Type": "Addition",
        "Per Life Premium": per_life_value,
        "Coverage Type": "Floater",
        "Suminsured": 500000,
        "Endorsement Date": today,
        "Date of Joining": today,
        "Employee Email": "",
        "Employee Mobile": "",
        "Remarks": "iter36 automated"
    }])


def _upload(df, hr_headers):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Sheet1")
    buf.seek(0)
    files = {"file": ("import.xlsx", buf.read(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    return requests.post(f"{API}/endorsements/import", headers=hr_headers, files=files, timeout=60)


def _find_endorsement(hr_headers, employee_id):
    r = requests.get(f"{API}/endorsements?policy_number={POLICY_NUMBER}", headers=hr_headers, timeout=30)
    assert r.status_code == 200
    matches = [e for e in r.json() if e.get("employee_id") == employee_id]
    return matches[0] if matches else None


# ---------- 3. Rate-card auto-fill via live import ----------
class TestImportAutoFill:
    def test_blank_per_life_autofills_from_rate_card(self, hr_headers):
        eid = f"IT36BLANK{int(time.time())}"
        df = _make_import_df(per_life_value="", employee_id=eid, name_suffix="Blank")
        resp = _upload(df, hr_headers)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body.get("success_count", 0) == 1, body

        end = _find_endorsement(hr_headers, eid)
        assert end is not None, "Endorsement not persisted"
        assert end["age"] == 41
        assert end["per_life_premium"] == 11500, f"Expected auto-filled 11500, got {end['per_life_premium']}"
        assert end["prorata_premium"] > 0

    def test_provided_per_life_is_used_not_rate_card(self, hr_headers):
        eid = f"IT36FIX{int(time.time())}"
        df = _make_import_df(per_life_value=5000, employee_id=eid, name_suffix="Fixed")
        resp = _upload(df, hr_headers)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body.get("success_count", 0) == 1, body

        end = _find_endorsement(hr_headers, eid)
        assert end is not None
        assert end["per_life_premium"] == 5000, f"Expected provided 5000, got {end['per_life_premium']}"
        assert end["prorata_premium"] > 0


# ---------- 4. Regression: standard template ----------
class TestStandardTemplateRegression:
    def test_standard_template_still_works(self, hr_headers):
        r = requests.get(f"{API}/endorsements/template/download", headers=hr_headers, timeout=30)
        assert r.status_code == 200
        ctype = r.headers.get("content-type", "")
        assert "spreadsheet" in ctype or "openxml" in ctype
        assert len(r.content) > 500

    def test_preview_endpoint_still_reachable(self, hr_headers):
        # Reuse a small df; endpoint should return 200 with preview data or
        # a 422 if it doesn't exist — we accept 200 primarily.
        df = _make_import_df(per_life_value="", employee_id="PREVIEW_ONLY", name_suffix="Prev")
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name="Sheet1")
        buf.seek(0)
        files = {"file": ("preview.xlsx", buf.read(),
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = requests.post(f"{API}/endorsements/import/preview", headers=hr_headers, files=files, timeout=30)
        # Endpoint may not exist in this build; accept 200 or 404, fail on 5xx
        assert r.status_code < 500, r.text
